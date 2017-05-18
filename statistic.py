#!/usr/local/bin/python3.5
from matplotlib import pyplot as plt
import openpyxl
import pymysql
from pylab import mpl
db = pymysql.connect('23.83.235.63', 'markdev', 'mark2017', 'codeclass', charset="utf8")
cur = db.cursor()

mpl.rcParams['font.sans-serif'] = ['FangSong']      # 指定默认字体
mpl.rcParams['axes.unicode_minus'] = False           # 解决保存图像是负号'-'显示为方块的问题
zjs = 3                                                 # 助教人数，确保助教先进群！


def get_times(lst: list, step: int):
    """
    对传入的列表 lst 按间隔 step 统计次数，返回字典
    :param lst:
    :param step:
    :return:
    """
    result = {}
    for i in lst:
        k = i//step
        if k in result:
            result[k] += 1
        else:
            result[k] = 1
    return result


def get_members(file):
    """
    从成员信息文件获取并返回全体成员用户名生成器
    """
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    # 从统计表格中第一位学员开始，3 = 2（题头）+ 1（从1开始计数）
    for row in ws[3+zjs:ws.max_row]:
        if row[0].value is None:
            yield row[1].value
        else:
            yield row[0].value


def _get_data(username):
    """
    从数据库获取并返回当前学员进度信息
    课程章节、课程总数、用户名
    """
    cur.execute('''
     SELECT MAX(chapter.seq), COUNT(lesson.seq), user.username
     FROM school_learnedlesson learned
     LEFT JOIN school_lesson lesson
     ON lesson.id = learned.lesson_id
     LEFT JOIN auth_user user
     ON user.id = learned.user_id
     LEFT JOIN school_chapter chapter
     ON chapter.id = lesson.chapter_id
     WHERE user.username = '{username}'
     GROUP BY user.username
     ;'''.format(username=username))
    return cur.fetchone()


def get_process(file):
    """
     获取全体学员进度信息，已知学员、未知学员
    """
    # process = {}
    process = []
    total = []
    members = []
    missing = []
    for m in get_members(file):
        data = _get_data(m)
        if data:
            # if data[0] in process:
            #     process[data[0]] += 1
            # else:
            #     process[data[0]] = 1
            process.append(data[0])
            total.append(data[1])
            members.append(m)
        else:
            missing.append(m)
    return process, total, members, missing


def plot_process(ax, process, limit):
    """
    课程进度图
    """
    limit += 0.5
    # x = list(process)
    # x.sort()
    # y = [process[a] for a in x]
    # ymax = max(y)
    # ylab = ['第%d章' % a for a in range(13)]
    # ylab.insert(0, ' ')

    process_count = get_times(process, 1)
    x = list(process_count.keys())
    y = list(process_count.values())
    # [print(i) for i in zip(x, y)]
    ymax = max(y)
    ylab = ['第%d章' % a for a in range(13)]

    # fig, ax = plt.subplots()

    ax.barh(x, y, color='blue')
    ax.set_title('课程进度统计（图一）', size=18)
    ax.spines['top'].set_color('none')
    ax.spines['right'].set_color('none')
    ax.spines['bottom'].set_color('none')
    ax.set_yticks(range(len(ylab)))
    ax.set_yticklabels(ylab)
    ax.set_xticks(range(ymax + 1))
    # ax.set_xlabel('人数', size=15)
    ax.set_ylabel('章节', size=15)
    ax.annotate('最低限度', (3, limit), (3, limit + 2), arrowprops=dict(facecolor='red'), size=14)
    ax.axhline(limit, color='r')

    # plt.title('课程进度统计')
    # x = list(process)
    # y = [process[i] for i in x]
    # plt.barh(x, y)
    # plt.axhline(limit, color='r')
    # plt.yticks(['a', 'b', 'c', 'd'])


def plot_total(ax, total, limit):
    """
    课程总数图
    """
    limit = limit//10 + 0.5
    total_count = get_times(total, 10)
    x = list(total_count.keys())
    y = list(total_count.values())
    ymax = max(y)
    ylab = ['%d课' % a for a in range(0, 100, 10)]

    ax.barh(x, y, color='green')
    ax.set_title('课程总数统计（图二）', size=18)
    ax.spines['top'].set_color('none')
    ax.spines['right'].set_color('none')
    ax.set_yticks(range(len(ylab)))
    ax.set_yticklabels(ylab)
    ax.set_xticks(range(ymax + 1))
    ax.set_ylabel('课程数', size=15)
    ax.set_xlabel('人数', size=15)
    ax.axhline(limit, color='r')
    ax.annotate('最低限度', (3, limit), (3, limit+2), arrowprops=dict(facecolor='red'), size=14)


def plot_members(members, missing):
    """
    成员信息饼状图
    """
    plt.title('成员统计')
    x = (len(members), len(missing))
    explode = (0, 0.1)
    labels = ('已统计人数%d' % x[0], '未统计人数%d' % x[1])
    autopct = '%3.1f%%'
    startangle = 90
    plt.pie(x, explode, labels, autopct=autopct, startangle=startangle)


def run(process_limit, total_limit, filename='members.xlsx', img='statistic.png'):
    process, total, members, missing = get_process(filename)
    fig, (ax1, ax2) = plt.subplots(2, 1, sharex='col')
    # plt.subplot(221)
    plot_process(ax1, process, process_limit)
    # plt.subplot(222)
    plot_total(ax2, total, total_limit)
    # plt.subplot(223)
    # plot_members(members, missing)
    plt.savefig(img)
    plt.show()


if __name__ == '__main__':
    # members_file = input('请输入群成员文件名：')
    run(7, 40, filename='20170424.xlsx')
