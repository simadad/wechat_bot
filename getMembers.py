import itchat
import openpyxl
title = ['群昵称', '昵称', '进度', 'UserName']


def get_members(name: str):
    """
    根据群名称，返回群ID、群成员生成器
    """
    rooms = itchat.search_chatrooms(name=name)
    username = rooms[0]['UserName']
    nickname = rooms[0]['NickName']
    # print(username)
    yield name, nickname, username
    room_info = itchat.update_chatroom(userName=username)
    members_info = room_info['MemberList']
    for m in members_info:
        yield m['DisplayName'], m['NickName'], m['UserName']


def save_members(members: get_members, filename: str) -> None:
    """
    接收成员列表和文件名，储存或更新成员信息
    """
    filename += '.xlsx'
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
    except FileNotFoundError as e:
        print(e)
        print('new file')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=2, column=3).value = '1;0'
    row = 2
    for column in range(1, 5):
        ws.cell(row=1, column=column).value = title[column-1]
    for name in members:
        ws.cell(row=row, column=1).value = name[0]
        ws.cell(row=row, column=2).value = name[1]
        ws.cell(row=row, column=4).value = name[2]
        row += 1
    wb.save(filename)


def run(unique_name):
    members = get_members(unique_name)
    save_members(members, unique_name)

if __name__ == '__main__':
    run('成长群')
