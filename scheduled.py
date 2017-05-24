#!/usr/local/bin/python3.5
import openpyxl
import getMembers
from mybot import send_text


def get_now_info(name):
    """
    传入群识别关键字，返回群UserName，当前进度, 统计表
    """
    file_name = name + '.xlsx'
    # try:
    #    wb = openpyxl.load_workbook(file_name)
    # except FileNotFoundError:
    getMembers.run(name)
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    username = ws.cell(row=2, column=4).value
    schedule_info = ws.cell(row=2, column=3).value
    now_schedule_row = schedule_info.split(';')[0]
    print('now_schedule_row: ', now_schedule_row)
    return username, int(now_schedule_row), wb


def get_new_schedule(now_schedule_row, file):
    """
    接收当前进度的行数，返回新进度、进度描述
    """
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    new_schedule = ws.cell(row=now_schedule_row+1, column=1).value
    pre_info = ws.cell(row=now_schedule_row+1, column=2).value
    return new_schedule, pre_info


def save_new_schedule(unique_name, now_schedule_row, new_schedule, wb_member):
    """
    更新，保存群进度信息
    """
    file_name = unique_name + '.xlsx'
    new_info = str(now_schedule_row+1) + ';' + new_schedule
    ws = wb_member.active
    ws.cell(row=2, column=3).value = new_info
    print(wb_member, file_name, new_info, 22222)
    wb_member.save(file_name)


def package_msg(new_schedule, pre_info):
    """
    清洗进度信息
    """
    if pre_info:
        washed_info = pre_info.replace(';', '\n')
        msg = '今日进度：\n' + new_schedule + '\n' + '主要任务：\n' + washed_info
    else:
        msg = False
    return msg


def run(unique_name, file='scheduled.xlsx'):
    username, now_schedule_row, wb_members = get_now_info(unique_name)
    new_schedule, pre_info = get_new_schedule(now_schedule_row, file)
    save_new_schedule(unique_name, now_schedule_row, new_schedule, wb_members)
    msg = package_msg(new_schedule, pre_info)
    if msg:
        send_text(msg, username)
    return msg

if __name__ == '__main__':
    file_scheduled = 'scheduled.xlsx'
    run('20170424', file_scheduled)
